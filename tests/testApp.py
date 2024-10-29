import pytest
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
import openpyxl
import os
import sys
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '../src')))


@pytest.fixture
def driver():
    chrome_options = Options()
    chrome_options.add_argument("--headless")  # Adiciona o modo headless
    chrome_options.add_argument("--no-sandbox")  # Necessário para contornar restrições de sandbox
    chrome_options.add_argument("--disable-dev-shm-usage")  # Ajuda a prevenir problemas de memória
    driver = webdriver.Chrome(options=chrome_options)
    yield driver
    driver.quit() 

def test_page_load(driver):
    driver.get('https://www.novaliderinformatica.com.br/computadores')
    assert "Computadores" in driver.title

def test_elements_found(driver):
    driver.get('https://www.novaliderinformatica.com.br/computadores')
    titulos = driver.find_elements(By.XPATH, "//div[@class='product-name']")
    precos = driver.find_elements(By.XPATH, "//span[@class='price-off']")
    
    assert len(titulos) > 0, "Nenhum título encontrado."
    assert len(precos) > 0, "Nenhum preço encontrado."

def test_excel_creation():
    # Cria uma planilha de teste
    workbook = openpyxl.Workbook()
    workbook.create_sheet('produtos')
    sheet_produtos = workbook['produtos']
    sheet_produtos['A1'].value = 'Produto'
    sheet_produtos['B1'].value = 'Preço'
    
    # Salva e verifica se o arquivo foi criado
    workbook.save('produtos_teste.xlsx')
    assert os.path.exists('produtos_teste.xlsx')
    
    # Limpeza do arquivo após o teste
    os.remove('produtos_teste.xlsx')

def test_data_entry_in_excel():
    workbook = openpyxl.Workbook()
    sheet_produtos = workbook.create_sheet('produtos')
    sheet_produtos['A1'] = 'Produto'
    sheet_produtos['B1'] = 'Preço'
    sheet_produtos.append(["Produto Teste", "R$ 199,99"])
    
    workbook.save('produtos_teste.xlsx')
    workbook_loaded = openpyxl.load_workbook('produtos_teste.xlsx')
    sheet_loaded = workbook_loaded['produtos']
    
    # Verifica se os dados foram inseridos corretamente
    assert sheet_loaded['A2'].value == "Produto Teste"
    assert sheet_loaded['B2'].value == "R$ 199,99"
    
    os.remove('produtos_teste.xlsx')
